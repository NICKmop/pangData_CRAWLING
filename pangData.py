from selenium import webdriver
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
from tkinter import messagebox
import openpyxl ,re ,logging, logging.handlers, pymysql

log_today = datetime.now();
log_today = str(log_today).split(" ")[0];

# path = "excel/2022년_TFO 장비온라인모니터링.xlsx";
path = "Z:/# ITHELPDESK/3. EAP 모니터링/2022년_TFO 장비온라인모니터링.xlsx";

log = logging.getLogger("Z:/# Util/ExcelPang/pangData_"+str(log_today)+".log");
log.setLevel(logging.DEBUG);

fileHandler = logging.FileHandler("Z:/# Util/ExcelPang/pangData_"+str(log_today)+".log");
streamHadler = logging.StreamHandler();

log.addHandler(fileHandler);
log.addHandler(streamHadler);

def driver1(driver):
    driver.get('http://10.12.1.27/#/dashboard/0oKI');
    html = driver.page_source;
    return html;
    
def driver2(driver):
    driver.get('http://10.12.1.27/#/dashboard/woRs');
    html = driver.page_source;
    return html;
    
def connect():
    driver = webdriver.Chrome('Z:/# Util/ExcelPang/chromeDriver/chromedriver.exe');
    driver.implicitly_wait(15);
    driver.get('http://10.12.1.27/pa/login');
    #계정정보 GET
    driver.find_element_by_name('username').send_keys('tfo-admin');
    driver.find_element_by_name('password').send_keys('123456');
    driver.find_element_by_css_selector('.md-raised.md-primary.md-button.md-default-theme').click();

    # return 값
    soup = BeautifulSoup(driver1(driver), 'html.parser');
    soup2 = BeautifulSoup(driver2(driver), 'html.parser');

    return soup, soup2;
def dbSqlExcute():
    log.info("ss");


#텍스트 값에 변수 등록 필요
listBox = [];
listTextBox = [];

# soup = con.connect();
soup = connect();

ulTotal = soup[0].select_one('#dashboardContentDiv > span > span > ul');
liTotal = ulTotal.select('li');

ulTotal2 = soup[1].select_one('#dashboardContentDiv > span > span > ul');
liTotal2 = ulTotal2.select('li');

for i in liTotal:
    listBox.append(i);

for i in liTotal2:
    listBox.append(i);

for j in range(0,len(listBox)):
    liText = listBox[j].text;
    listTextBox.append(liText);

def replFromat(value,formatString1,formatString2):
    value = value.replace(formatString1,formatString2);
    value = value.replace(" ","");
    return value;


def reSplit(value):
    equipTime = [];
    equipName = [];
    RiChange = [];
    result = [];
    # 설비장비 | 일자 GET
    try:
        for i in value:
            i = i.strip();


            Ri = i.split("     ")[0]; # 설비명
            Li = i.split("     ")[1]; # 날짜확인

            if "LVAD" in Ri:
                RiChange = replFromat(Ri,"VA","-VA");
                Ri = RiChange;
            elif "JVAD" in Ri:
                RiChange = replFromat(Ri,"VA","-VA");
                Ri = RiChange;
            elif "O" in Ri:
                RiChange = replFromat(Ri,"VA","-VA");
                Ri = RiChange;
            elif "LATH" in Ri:
                RiChange = replFromat(Ri,"LATH","L-Lathe");
                Ri = RiChange;
            elif "LFUR" in Ri:
                RiChange = replFromat(Ri,"LFUR","L-Furnace");
                Ri = RiChange;
            elif "VFUR" in Ri:
                RiChange = replFromat(Ri,"VFUR","V-Furnace");
                Ri = RiChange;
            elif "RFUR" in Ri:
                RiChange = replFromat(Ri,"RFUR","R-Furnace");
                Ri = RiChange;
            elif "Tapering" in Ri:
                RiChange = replFromat(Ri, "Tapering", "Tapering1");
                Ri = RiChange;
            elif "Sintering" in Ri:
                RiChange = replFromat(Ri, "Sintering", "Sintering1");
                Ri = RiChange;

            if "VLATH" in Ri:
                RiChange = replFromat(Ri,"VLATH","V-Lathe");
                Ri = RiChange;
            elif "LDRAW" in Ri:
                RiChange = replFromat(Ri, "LDRAW","L-Drawing-");
                Ri = RiChange;
            elif "TDRAW" in Ri:
                RiChange = replFromat(Ri, "TDRAW","T-Drawing-");
                Ri = RiChange;
            elif "NDRAW" in Ri:
                RiChange = replFromat(Ri, "NDRAW","N-Drawing-");
                Ri = RiChange;
            elif "REW" in Ri:
                RiChange = replFromat(Ri, "REW","Rewinding");
                Ri = RiChange;

            equipTime.append(Li);
            equipName.append(Ri);
    except IndexError as iOn:
        print(iOn);

    for i in zip(equipName, equipTime):
        result.append(i);

    return result;

def log_output(eqipNameExcel, data):
    log.info('{} : {}'.format(str(eqipNameExcel),str(data)));

def todayCell(Today, load_ws,data,load_ws_Add_Data,yesterday):
    for cols in load_ws.iter_cols(min_row=6):
        cellAlphabet = str(cols[1]).split(".")[1];
        #어제자 제외 및 없음 등 추출 하기 위한
        if str(cols[1].value).split(" ")[0].replace("-",".") == Today:
            sliceTodayFull = cellAlphabet;
            if len(sliceTodayFull) == 4:
                sliceToday = cellAlphabet[0:2];
            else:
                sliceToday = cellAlphabet[0:1];

            #알파벳 변수로 받고 none / 제외 그대로
            for i in range(7,len(data)+20):
                # 설비명
                eqipNameExcel = load_ws['E'+str(i)].value;
                # 장비 현황
                eqipNameStatus = load_ws['F'+str(i)].value;
                for j in range(7,len(data)):
                    if eqipNameExcel == data[j][0]:
                        if data[j][1] == "몇초 전":
                            load_ws[sliceToday+str(i)].value = "";
                            log_output(eqipNameExcel, data[j][1]);
                            
                        elif "시간 전" in data[j][1]:
                            load_ws[sliceToday+str(i)].value = "";
                            log_output(eqipNameExcel, data[j][1]);
                            
                        elif "분 전" in data[j][1]:
                            load_ws[sliceToday+str(i)].value = "";
                            log_output(eqipNameExcel, data[j][1]);
                            
                        elif "하루 전" in data[j][1]:
                            load_ws[sliceToday+str(i)].value = "";
                            log_output(eqipNameExcel, data[j][1]);
                           
                        elif "2일 전" in data[j][1]:
                            load_ws[sliceToday+str(i)].value = "";
                            log_output(eqipNameExcel, data[j][1]);
                            
                        elif "3일 전" in data[j][1]:
                            load_ws[sliceToday+str(i)].value = "";
                            log_output(eqipNameExcel, data[j][1]);
                           
                        elif "4일 전" in data[j][1]:
                            load_ws[sliceToday+str(i)].value = "";
                            log_output(eqipNameExcel, data[j][1]);
                           
                        elif "5일 전" in data[j][1]:
                            load_ws[sliceToday+str(i)].value = "";
                            log_output(eqipNameExcel, data[j][1]);
                            
                        else:
                            if eqipNameStatus:
                                load_ws[sliceToday+str(i)].value = "";
                                
                            else:
                                load_ws[sliceToday+str(i)].value = data[j][1];
                                log_output(eqipNameExcel, data[j][1]);
                               

# 어제자 날짜 구하기
def yesterDayCell(yesterday,load_ws,data,wordBox,weekDatetime):
    if weekDatetime == '6':
        yesterday = datetime.strptime(yesterday,"%Y.%m.%d") - timedelta(days=+2);
        yesterday = str(yesterday).split(" ")[0];

    for cols in load_ws.iter_cols(min_row=6):
        cellAlphabet = str(cols[1]).split(".")[1];
        if str(cols[1].value).split(" ")[0].replace("-",".") == yesterday:
            sliceYesterdayFull = cellAlphabet;
            if len(sliceYesterdayFull) == 4:
                sliceYesterday =  cellAlphabet[0:2];
            else:
                sliceYesterday =  cellAlphabet[0:1];
           
            sliceToday = wordBox[0];
            for i in range(7, len(data) + 20):
                # 함수화 가능 할듯.
                yesDt = load_ws[sliceYesterday+str(i)].value
                if load_ws[sliceYesterday+str(i)].value == "제외":
                    #셀 번호 확인
                    cellNumber = str(load_ws[sliceYesterday+str(i)]).split(".")[1].replace(">","")
                    load_ws[sliceToday+str(i)].value = yesDt;
                elif load_ws[sliceYesterday+str(i)].value == "없음":
                    cellNumber = str(load_ws[sliceYesterday+str(i)]).split(".")[1].replace(">","")
                    load_ws[sliceToday+str(i)].value = yesDt;
                elif load_ws[sliceYesterday+str(i)].value == "보류":
                    cellNumber = str(load_ws[sliceYesterday+str(i)]).split(".")[1].replace(">","")
                    load_ws[sliceToday+str(i)].value = yesDt;

# 엑셀 쓰기
def excelReadWrite(path, value):
    data = reSplit(value);
    #더미 값 입력
    for i in zip("0", "0"):
        data.insert(0,i);
        data.insert(1,i);
        data.insert(2,i);
        data.insert(3,i);
        data.insert(4,i);
        data.insert(5,i);
        data.insert(6,i);
        data.insert(7,i);
    month = datetime.now().month;
    # 현재 날짜 데이터 관련
    Today = datetime.now();
    yesterday = Today - timedelta(1);

    Today = str(Today).split(" ")[0].replace("-",".");
    yesterday = str(yesterday).split(" ")[0].replace("-",".");
    wordBox = [];
    
    # 시트 이름 지정
    sheetName = str(month)+'월_일일점검';
    sheetAddName = '통계';

    datetime_date = datetime.strptime(Today, '%Y.%m.%d');
    # 0:월 1:화 2:수 3:목 4:금 5:토 6:일
    weekDatetime = datetime_date.weekday();

    if weekDatetime == 5:
        exit();
    elif weekDatetime == 6:
        exit();

    load_monitoring = openpyxl.load_workbook(path);
    load_ws = load_monitoring[sheetName];
    load_ws_Add_Data = load_monitoring[sheetAddName];

    for cols in load_ws.iter_cols(min_row=6):
        cellAlphabet = str(cols[1]).split(".")[1];
        #어제자 제외 및 없음 등 추출 하기 위한
        if str(cols[1].value).split(" ")[0].replace("-",".") == Today:
            sliceTodayFull = cellAlphabet;
            if len(sliceTodayFull) == 4:
                sliceToday = cellAlphabet[0:2];
            else:
                sliceToday = cellAlphabet[0:1];
            wordBox.append(sliceToday);

    todayCell(Today,load_ws,data,load_ws_Add_Data,yesterday);
    yesterDayCell(yesterday,load_ws,data,wordBox,weekDatetime);

    load_monitoring.save(path);

    messagebox.showinfo("확인 창", "팡데이터 엑셀 저장 완료");


excelReadWrite(path,listTextBox);

# exl.excelCheckopen(path);

